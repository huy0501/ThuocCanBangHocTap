import uuid
import unicodedata
import importlib.util
from datetime import datetime, date, time as dt_time

import streamlit as st
import streamlit.components.v1 as components
import matplotlib.pyplot as plt


# =========================
# Page config
# =========================
st.set_page_config(
    page_title="Bảng Điều Khiển - Thước Cân Bằng",
    layout="wide",
    initial_sidebar_state="collapsed",
)

# =========================
# CSS
# =========================
CSS = """
<style>
#MainMenu {visibility:hidden;}
footer {visibility:hidden;}
header {visibility:hidden;}

section[data-testid="stSidebar"] { display: none !important; }
div[data-testid="collapsedControl"] { visibility: hidden !important; width: 0 !important; height: 0 !important; }

[data-testid="stAppViewContainer"] { background:#f5f6f8; }
.block-container { padding-top: 1.1rem; padding-bottom: 6rem; }
h1 { font-weight: 900; letter-spacing: -0.02em; }

/* LEFT NAV */
.leftNav {
  background: #0b0f19;
  color: #e5e7eb;
  border-radius: 14px;
  padding: 14px 12px;
  height: calc(100vh - 2.2rem);
  position: sticky;
  top: 1.1rem;
  overflow: auto;
  border: 1px solid rgba(255,255,255,0.06);
}
.leftNav .brand { font-weight: 900; font-size: 18px; margin-bottom: 12px; }
.leftNav .navItem {
  background: rgba(255,255,255,0.06);
  border: 1px solid rgba(255,255,255,0.06);
  border-radius: 12px;
  padding: 10px 10px;
  margin: 8px 0;
  font-weight: 800;
}
.leftNav .navItem.muted { background: transparent; border: 1px solid rgba(255,255,255,0.06); }
.leftNav small { opacity: 0.85; font-weight: 600; }

/* SUBJECT CARDS */
.course-card {
  border-radius: 12px;
  overflow: hidden;
  background: #fff;
  border: 1px solid rgba(17,24,39,0.10);
  box-shadow: 0 1px 0 rgba(17,24,39,0.04);
}
.course-top { height: 120px; position: relative; }
.course-dots {
  position:absolute;
  right: 10px; top: 10px;
  color: rgba(255,255,255,0.95);
  font-size: 22px;
  font-weight: 900;
}
.course-body { padding: 12px 14px 10px 14px; }
.course-slug { font-weight: 900; font-size: 16px; margin-bottom: 2px; text-transform: lowercase; }
.course-code { color:#6b7280; font-size: 13px; line-height: 1.25; }
.course-icons { margin-top: 10px; color:#6b7280; font-size: 16px; }
.course-icons span { margin-right: 14px; }

/* RIGHT PANEL */
.rightPanel {
  background: #ffffff;
  border: 1px solid rgba(17,24,39,0.10);
  border-radius: 14px;
  padding: 14px 14px;
  box-shadow: 0 1px 0 rgba(17,24,39,0.04);
  position: sticky;
  top: 1.1rem;
  max-height: calc(100vh - 2.2rem);
  overflow: auto;
}
.todo-title { font-size: 18px; font-weight: 900; margin: 0 0 10px 0; }
.todo-item {
  display: grid;
  grid-template-columns: 18px 1fr 14px;
  gap: 10px;
  padding: 10px 10px;
  border-radius: 12px;
  border: 1px solid rgba(17,24,39,0.08);
  background: #fafafa;
  margin-bottom: 10px;
}
.todo-item b { display:block; font-size: 13px; line-height: 1.25; }
.todo-item small { color:#6b7280; display:block; margin-top: 4px; }
.todo-x { color:#9ca3af; font-weight: 900; }

/* MID PANEL */
.midPanel {
  background: #ffffff;
  border: 1px solid rgba(17,24,39,0.10);
  border-radius: 14px;
  padding: 14px 14px;
  box-shadow: 0 1px 0 rgba(17,24,39,0.04);
  margin-top: 14px;
}

/* Hide widget labels */
label[data-testid="stWidgetLabel"] { display:none !important; }

/* CHATBOX UI */
.chat-shell {
  background: #fff;
  border: 1px solid rgba(17,24,39,0.12);
  border-radius: 16px;
  box-shadow: 0 18px 40px rgba(0,0,0,0.20);
  overflow: hidden;
}
.chat-head {
  background: #0b0f19;
  color: #e5e7eb;
  padding: 10px 12px;
  display:flex;
  justify-content: space-between;
  align-items:center;
  font-weight: 900;
}
.chat-hint { font-size: 12px; opacity: .9; font-weight: 800; }
.chat-body {
  background: #f8fafc;
  padding: 10px 10px;
  overflow-y: auto;
}
.bubble {
  background: #fff;
  border: 1px solid rgba(17,24,39,0.08);
  border-radius: 12px;
  padding: 8px 10px;
  margin: 8px 0;
}
.bubble.user { border-left: 4px solid #2563eb; }
.bubble.ai { border-left: 4px solid #10b981; }
.bname { font-weight: 900; margin-bottom: 2px; }

.chat-foot {
  background: #fff;
  border-top: 1px solid rgba(17,24,39,0.10);
  padding: 10px 10px;
}

/* Minimized pill */
.chat-pill {
  background: #2563eb;
  color: #fff;
  border-radius: 999px;
  padding: 10px 14px;
  display:flex;
  align-items:center;
  gap: 10px;
  box-shadow: 0 12px 28px rgba(0,0,0,0.18);
  border: 1px solid rgba(17,24,39,0.12);
  font-weight: 900;
}
.chat-pill .dot {
  width: 10px; height: 10px; border-radius: 999px;
  background: rgba(255,255,255,0.9);
}
.chat-pill .miniBtn button {
  height: 34px !important;
  border-radius: 999px !important;
  background: rgba(255,255,255,0.15) !important;
  color: #fff !important;
  border: 1px solid rgba(255,255,255,0.22) !important;
  font-weight: 900 !important;
  padding: 0 12px !important;
}
.chat-pill .miniBtn button:hover { filter: brightness(1.05); }

@media (max-width: 520px) {
  .block-container { padding-bottom: 9rem; }
}
</style>
"""
st.markdown(CSS, unsafe_allow_html=True)


# =========================
# Floating helper (no streamlit-float)
# =========================
def st_float_container(key: str, css: str):
    anchor_id = f"float-anchor-{key}-{uuid.uuid4().hex}"
    st.markdown(f'<div id="{anchor_id}"></div>', unsafe_allow_html=True)

    components.html(
        f"""
        <script>
        (function() {{
          const run = () => {{
            const doc = window.parent.document;
            const anchor = doc.getElementById("{anchor_id}");
            if (!anchor) return;

            let target =
              anchor.closest('div[data-testid="stVerticalBlock"]') ||
              anchor.closest('div[data-testid="stHorizontalBlock"]');

            if (!target) {{
              let el = anchor;
              for (let i = 0; i < 60; i++) {{
                if (!el) break;
                const dt = el.getAttribute && el.getAttribute("data-testid");
                if (dt === "stVerticalBlock" || dt === "stHorizontalBlock") {{
                  target = el;
                  break;
                }}
                el = el.parentElement;
              }}
            }}

            if (!target) return;

            target.style.cssText += `{css}`;
            target.style.zIndex = "999999";
            target.style.pointerEvents = "auto";
            target.style.margin = "0";
          }};

          setTimeout(run, 50);
          setTimeout(run, 250);
        }})();
        </script>
        """,
        height=0,
        width=0,
    )


# =========================
# Data
# =========================
SUBJECT_CARDS = [
    {"name": "Toán",       "slug": "math",       "code": "THCS.GP-MATH-8", "year": "2025-2026", "color": "#1f77b4"},
    {"name": "Khoa học",   "slug": "science",    "code": "THCS.GP-SCI-8",  "year": "2025-2026", "color": "#8e44ad"},
    {"name": "Tin học",    "slug": "computer",   "code": "THCS.GP-IT-8",   "year": "2025-2026", "color": "#9a7421"},
    {"name": "Tiếng Anh",  "slug": "english",    "code": "THCS.GP-ENG-8",  "year": "2025-2026", "color": "#0f7a2b"},
    {"name": "Văn",        "slug": "literature", "code": "THCS.GP-VNS-8",  "year": "2025-2026", "color": "#ff5c00"},
    {"name": "Lịch Sử",    "slug": "history",    "code": "THCS.GP-HIS-8",  "year": "2025-2026", "color": "#2c3e50"},
]

EXERCISES_DEFAULT = {
    "Văn": [
        "Đọc 1 đoạn 250–400 chữ, gạch chân 5 từ khóa và viết 3 câu tóm tắt.",
        "Viết 1 đoạn văn 120–150 chữ theo cấu trúc: Mở–Thân–Kết (1 ý chính).",
        "Làm 5 câu hỏi đọc hiểu (ý chính, chi tiết, suy luận, từ vựng, thông điệp).",
    ],
    "Anh": [
        "Ôn 15 từ vựng theo chủ đề + viết 5 câu dùng từ mới.",
        "Làm 10 câu ngữ pháp (thì/so sánh/mệnh đề quan hệ) + chữa lỗi.",
        "Nghe 3–5 phút (video ngắn) rồi ghi 5 ý chính bằng tiếng Anh.",
    ],
    "Sử": [
        "Vẽ timeline 6 mốc sự kiện của 1 bài học + 1 câu giải thích/mốc.",
        "Làm 8 câu trắc nghiệm + ghi lại 3 câu sai và lý do sai.",
        "Viết 5 flashcards: Sự kiện – Nguyên nhân – Kết quả.",
    ],
    "Toán": [
        "Làm 10 bài mức cơ bản theo 1 dạng + ghi 2 lỗi hay gặp.",
        "Làm 5 bài vận dụng nhẹ + viết 1 lời giải mẫu ngắn gọn.",
        "Chữa 3 bài sai gần nhất: ghi 'sai ở đâu – sửa thế nào'.",
    ],
    "Khoa học": [
        "Làm 8 câu hỏi khái niệm + 2 bài tính đơn giản (nếu có).",
        "Tóm tắt 1 trang kiến thức bằng sơ đồ (mindmap) ~10 nhánh.",
        "Quan sát/thí nghiệm mô phỏng đơn giản và viết 5 dòng kết luận.",
    ],
    "Tin": [
        "Làm 1 bài thuật toán cơ bản (điều kiện/vòng lặp) và thử 3 test case.",
        "Sửa 1 đoạn code lỗi (tự tạo hoặc bài cũ): ghi lỗi và cách sửa.",
        "Tạo 1 chương trình nhỏ 20–30 dòng (tính TB/ phân loại).",
    ],
}

PASS_THRESHOLD = 6.5
FOCUS_K = 3

ALL_SUBJECTS = ["Toán", "Khoa học", "Tin", "Văn", "Sử", "Anh"]
DISPLAY_NAME = {"Anh": "Tiếng Anh", "Sử": "Lịch Sử"}

SUBJ_KEY = {
    "Toán": "toan",
    "Khoa học": "khoa_hoc",
    "Tin": "tin",
    "Văn": "van",
    "Sử": "su",
    "Anh": "anh",
}

QUIZ_TYPE_MCQ = "Trắc nghiệm (MCQ)"
QUIZ_TYPE_SHORT = "Tự luận (Short answer)"
QUIZ_TYPES = [QUIZ_TYPE_MCQ, QUIZ_TYPE_SHORT]
MCQ_LETTERS = ["A", "B", "C", "D"]


# =========================
# Utils
# =========================
def has_openpyxl() -> bool:
    return importlib.util.find_spec("openpyxl") is not None

def normalize_text(text: str) -> str:
    t = (text or "").lower().strip()
    t = unicodedata.normalize("NFD", t)
    t = "".join(ch for ch in t if unicodedata.category(ch) != "Mn")
    t = "".join(ch if ch.isalnum() or ch.isspace() else " " for ch in t)
    return " ".join(t.split())

def parse_bool(v) -> bool:
    if v is None:
        return False
    if isinstance(v, bool):
        return v
    if isinstance(v, (int, float)):
        return v != 0
    s = normalize_text(str(v))
    return s in ["1", "true", "yes", "y", "co", "bat buoc", "mandatory", "required", "x"]

def clamp_int(v, lo: int, hi: int, default: int):
    try:
        x = int(v)
    except Exception:
        return default
    return max(lo, min(hi, x))

def safe_str(v) -> str:
    return "" if v is None else str(v)

def format_due(dt: datetime | None) -> str:
    if dt is None:
        return "Không deadline"
    return dt.strftime("%d/%m/%Y %H:%M")

def parse_due(value):
    """
    Accept:
    - datetime
    - date (-> 23:59)
    - string: 'YYYY-MM-DD HH:MM' OR ISO 'YYYY-MM-DDTHH:MM:SS'
    """
    if value is None:
        return None
    if isinstance(value, datetime):
        return value
    if isinstance(value, date) and not isinstance(value, datetime):
        return datetime(value.year, value.month, value.day, 23, 59)
    s = safe_str(value).strip()
    if not s:
        return None
    # try iso
    try:
        return datetime.fromisoformat(s)
    except Exception:
        pass
    # try common format
    for fmt in ["%Y-%m-%d %H:%M", "%Y-%m-%d %H:%M:%S", "%d/%m/%Y %H:%M", "%d/%m/%Y %H:%M:%S"]:
        try:
            return datetime.strptime(s, fmt)
        except Exception:
            continue
    # try only date
    for fmt in ["%Y-%m-%d", "%d/%m/%Y"]:
        try:
            d = datetime.strptime(s, fmt).date()
            return datetime(d.year, d.month, d.day, 23, 59)
        except Exception:
            continue
    return None

def compute_scores(row):
    math = float(row.get("Toán", 0))
    sci  = float(row.get("Khoa học", 0))
    it   = float(row.get("Tin", 0))
    lit  = float(row.get("Văn", 0))
    hist = float(row.get("Sử", 0))
    eng  = float(row.get("Anh", 0))

    logic_avg = round((math + sci + it) / 3, 2)
    lang_avg  = round((lit + hist + eng) / 3, 2)
    diff = round(abs(logic_avg - lang_avg), 2)

    if diff < 1:
        level = "Cân bằng tốt"
    elif diff < 2:
        level = "Lệch vừa"
    else:
        level = "Lệch rõ"
    return logic_avg, lang_avg, diff, level

def student_label(s):
    return f'{s["ID"]} | {s["Lớp"]} | {s["Tên"]}'

def weakest_subjects(scores_dict, group: str, k: int = 2):
    subs = ["Toán", "Khoa học", "Tin"] if group == "Logic" else ["Văn", "Sử", "Anh"]
    scored = sorted([(s, scores_dict[s]) for s in subs], key=lambda x: x[1])
    return [s for s, _ in scored[:k]]

def get_low_subjects(scores: dict, threshold: float = PASS_THRESHOLD):
    low = []
    for s in ALL_SUBJECTS:
        v = float(scores.get(s, 0))
        if v < threshold:
            low.append((s, v))
    low.sort(key=lambda x: x[1])
    return low

def html_safe(text: str) -> str:
    if text is None:
        return ""
    text = (text.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;"))
    return text.replace("\n", "<br>")


# =========================
# Quiz Bank + Excel
# =========================
def set_teacher_notice(subject: str, kind: str, msg: str):
    st.session_state.teacher_notice = {"subject": subject, "kind": kind, "msg": msg}

def set_import_notice(kind: str, msg: str, details: str | None = None):
    st.session_state.import_notice = {"kind": kind, "msg": msg, "details": details}

def get_quizzes(subject: str):
    bank = st.session_state.get("quiz_bank", {})
    return list(bank.get(subject, []))

def all_quizzes():
    out = []
    for s in ALL_SUBJECTS:
        out.extend(get_quizzes(s))
    return out

def find_quiz(subject: str, quiz_id: str):
    for q in get_quizzes(subject):
        if q["id"] == quiz_id:
            return q
    return None

def quiz_badge(q: dict) -> str:
    return "✅ BẮT BUỘC" if q.get("mandatory") else "⭕ TỰ CHỌN"

def quiz_task_title(q: dict) -> str:
    diff = int(q.get("difficulty", 3))
    pts = int(q.get("points", 10))
    return f"📝 Quiz {quiz_badge(q)} ⭐{diff}/5 ({pts}đ): {q.get('title','(No title)')}"

def make_template_xlsx_bytes():
    if not has_openpyxl():
        return None
    from openpyxl import Workbook
    from io import BytesIO

    wb = Workbook()
    ws = wb.active
    ws.title = "Quiz"

    headers = [
        "Subject", "Title", "Question", "Type",
        "A", "B", "C", "D",
        "Correct", "Mandatory", "Difficulty",
        "Points", "MaxAttempts", "DueAt",
        "Explanation", "AnswerText"
    ]
    ws.append(headers)

    ws.append([
        "Toán",
        "Quiz mẫu - PT bậc 1",
        "Giải phương trình: 2x + 3 = 7. x = ?",
        "MCQ",
        "1", "2", "3", "4",
        "B",
        1,
        2,
        10,
        2,
        "2026-02-13 23:59",
        "Chuyển vế: 2x = 4 => x = 2",
        ""
    ])

    ws.append([
        "Văn",
        "Quiz mẫu - Tự luận",
        "Viết 3 câu tóm tắt đoạn văn (80-100 chữ).",
        "SHORT",
        "", "", "", "",
        "",
        0,
        3,
        10,
        1,
        "",
        "Nhắc: có ý chính + chi tiết + kết luận.",
        "Đáp án mẫu: (giáo viên điền...)"
    ])

    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()

def export_quiz_bank_xlsx_bytes():
    if not has_openpyxl():
        return None
    from openpyxl import Workbook
    from io import BytesIO

    wb = Workbook()
    ws = wb.active
    ws.title = "Quiz"

    headers = [
        "Subject", "Title", "Question", "Type",
        "A", "B", "C", "D",
        "Correct", "Mandatory", "Difficulty",
        "Points", "MaxAttempts", "DueAt",
        "Explanation", "AnswerText"
    ]
    ws.append(headers)

    for q in all_quizzes():
        subject = q.get("subject", "")
        title = q.get("title", "")
        question = q.get("question", "")
        mandatory = 1 if q.get("mandatory") else 0
        difficulty = int(q.get("difficulty", 3))
        points = int(q.get("points", 10))
        max_attempts = int(q.get("max_attempts", 1))
        due_at = q.get("due_at") or ""
        explanation = q.get("explanation") or ""

        if q.get("type") == "mcq":
            opts = q.get("options", {}) or {}
            row = [
                subject, title, question, "MCQ",
                opts.get("A", ""), opts.get("B", ""), opts.get("C", ""), opts.get("D", ""),
                q.get("correct_letter", ""),
                mandatory, difficulty,
                points, max_attempts,
                due_at,
                explanation,
                ""
            ]
        else:
            row = [
                subject, title, question, "SHORT",
                "", "", "", "",
                "",
                mandatory, difficulty,
                points, max_attempts,
                due_at,
                explanation,
                q.get("answer_text", "")
            ]

        ws.append(row)

    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()

def normalize_header(h: str) -> str:
    return normalize_text(h).replace(" ", "").replace("_", "")

def map_subject(value):
    s = normalize_text(safe_str(value))
    if not s:
        return None
    if "toan" in s or s == "math":
        return "Toán"
    if "khoahoc" in s or "science" in s or s == "sci":
        return "Khoa học"
    if "tin" in s or s == "it" or "computer" in s or "program" in s:
        return "Tin"
    if "van" in s or "nguvan" in s or "literature" in s:
        return "Văn"
    if "su" in s or "lichsu" in s or "history" in s:
        return "Sử"
    if "anh" in s or "tienganh" in s or "english" in s:
        return "Anh"
    return None

def parse_quiz_type(value) -> str:
    s = normalize_text(safe_str(value))
    if any(k in s for k in ["mcq", "trac", "choice", "multiple"]):
        return "mcq"
    if any(k in s for k in ["short", "tu luan", "essay", "tuluan"]):
        return "short"
    return "mcq"

def parse_correct_letter(value):
    if value is None:
        return None
    if isinstance(value, (int, float)):
        idx = int(value)
        if 1 <= idx <= 4:
            return MCQ_LETTERS[idx - 1]
        return None
    s = safe_str(value).strip().upper()
    if not s:
        return None
    c = s[:1]
    if c in MCQ_LETTERS:
        return c
    if s in ["1", "2", "3", "4"]:
        return MCQ_LETTERS[int(s) - 1]
    return None

def import_quizzes_from_excel():
    up = st.session_state.get("import_excel_file", None)
    if up is None:
        set_import_notice("error", "Bạn chưa upload file Excel.")
        return
    if not has_openpyxl():
        set_import_notice("error", "Thiếu openpyxl. Cài: pip install openpyxl")
        return

    from openpyxl import load_workbook

    try:
        up.seek(0)
        wb = load_workbook(up, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
    except Exception as e:
        set_import_notice("error", f"Không đọc được file Excel: {e}")
        return

    if not rows or len(rows) < 2:
        set_import_notice("error", "File Excel không có dữ liệu (cần header + ít nhất 1 dòng quiz).")
        return

    headers_raw = rows[0]
    headers = [safe_str(h).strip() for h in headers_raw]
    hmap = {normalize_header(h): i for i, h in enumerate(headers) if safe_str(h).strip()}

    def cell(row, *keys):
        for k in keys:
            idx = hmap.get(normalize_header(k))
            if idx is not None and idx < len(row):
                return row[idx]
        return None

    imported = 0
    failed = 0
    fail_msgs = []

    for r_idx, row in enumerate(rows[1:], start=2):
        if row is None or all(v is None or safe_str(v).strip() == "" for v in row):
            continue

        try:
            subject = map_subject(cell(row, "Subject", "Mon", "Môn", "Course"))
            if subject is None:
                raise ValueError(f"Dòng {r_idx}: Môn không hợp lệ.")

            title = cell(row, "Title", "Tieu de", "Tiêu đề", "Name")
            question = cell(row, "Question", "Cau hoi", "Câu hỏi", "Content")
            if safe_str(title).strip() == "":
                raise ValueError(f"Dòng {r_idx}: thiếu Title.")
            if safe_str(question).strip() == "":
                raise ValueError(f"Dòng {r_idx}: thiếu Question.")

            qtype = parse_quiz_type(cell(row, "Type", "Loai", "Loại"))

            mandatory = parse_bool(cell(row, "Mandatory", "Bat buoc", "Bắt buộc", "Required"))
            difficulty = clamp_int(cell(row, "Difficulty", "Do kho", "Độ khó", "Level"), 1, 5, 3)
            points = clamp_int(cell(row, "Points", "Diem", "Điểm"), 1, 1000, 10)
            max_attempts = clamp_int(cell(row, "MaxAttempts", "Max attempts", "Attempts"), 0, 999, 1)  # 0 = unlimited
            due = parse_due(cell(row, "DueAt", "Due", "Deadline", "Han nop", "Hạn nộp"))
            due_at = due.isoformat(timespec="seconds") if due else ""

            explanation = safe_str(cell(row, "Explanation", "Giai thich", "Giải thích", "Note")).strip()

            quiz = {
                "id": uuid.uuid4().hex,
                "subject": subject,
                "title": safe_str(title).strip(),
                "question": safe_str(question).strip(),
                "type": qtype,
                "mandatory": bool(mandatory),
                "difficulty": int(difficulty),
                "points": int(points),
                "max_attempts": int(max_attempts),
                "due_at": due_at if due_at else None,
                "explanation": explanation,
                "created_at": datetime.now().isoformat(timespec="seconds"),
            }

            if qtype == "mcq":
                options = {
                    "A": safe_str(cell(row, "A", "Option A")).strip(),
                    "B": safe_str(cell(row, "B", "Option B")).strip(),
                    "C": safe_str(cell(row, "C", "Option C")).strip(),
                    "D": safe_str(cell(row, "D", "Option D")).strip(),
                }
                non_empty = [L for L in MCQ_LETTERS if options.get(L)]
                if len(non_empty) < 2:
                    raise ValueError(f"Dòng {r_idx}: MCQ cần >= 2 lựa chọn A-D.")

                correct_letter = parse_correct_letter(cell(row, "Correct", "Dap an", "Đáp án", "Key"))
                if correct_letter is None:
                    raise ValueError(f"Dòng {r_idx}: Correct phải là A/B/C/D hoặc 1..4.")
                if not options.get(correct_letter):
                    raise ValueError(f"Dòng {r_idx}: Correct={correct_letter} nhưng lựa chọn trống.")

                quiz["options"] = options
                quiz["correct_letter"] = correct_letter

            else:
                ans = safe_str(cell(row, "AnswerText", "Answer", "Dap an mau", "Đáp án mẫu")).strip()
                if not ans:
                    raise ValueError(f"Dòng {r_idx}: SHORT cần AnswerText.")
                quiz["answer_text"] = ans

            bank = st.session_state.quiz_bank
            bank.setdefault(subject, [])
            bank[subject] = bank[subject] + [quiz]
            st.session_state.quiz_bank = bank
            imported += 1

        except Exception as e:
            failed += 1
            fail_msgs.append(str(e))

    if imported > 0 and failed == 0:
        set_import_notice("success", f"✅ Import thành công {imported} quiz.")
    elif imported > 0 and failed > 0:
        details = "\n".join(fail_msgs[:12]) + (("\n...") if len(fail_msgs) > 12 else "")
        set_import_notice("warning", f"⚠️ Import được {imported} quiz, lỗi {failed} dòng.", details=details)
    else:
        details = "\n".join(fail_msgs[:12]) + (("\n...") if len(fail_msgs) > 12 else "")
        set_import_notice("error", f"❌ Import thất bại. Lỗi {failed} dòng.", details=details)


# =========================
# Teacher create / edit quiz helpers
# =========================
def add_quiz_from_inputs(subject: str):
    sk = SUBJ_KEY[subject]

    title = (st.session_state.get(f"new_title_{sk}", "") or "").strip()
    question = (st.session_state.get(f"new_question_{sk}", "") or "").strip()
    qtype_ui = st.session_state.get(f"new_type_{sk}", QUIZ_TYPE_MCQ)
    mandatory = bool(st.session_state.get(f"new_mand_{sk}", False))
    difficulty = clamp_int(st.session_state.get(f"new_diff_{sk}", 3), 1, 5, 3)
    points = clamp_int(st.session_state.get(f"new_points_{sk}", 10), 1, 1000, 10)

    unlimited = bool(st.session_state.get(f"new_unlim_{sk}", False))
    max_attempts = 0 if unlimited else clamp_int(st.session_state.get(f"new_attempts_{sk}", 1), 1, 99, 1)

    has_due = bool(st.session_state.get(f"new_has_due_{sk}", False))
    due_at = None
    if has_due:
        d = st.session_state.get(f"new_due_date_{sk}", None)
        t = st.session_state.get(f"new_due_time_{sk}", None)
        if isinstance(d, date) and isinstance(t, dt_time):
            due_dt = datetime(d.year, d.month, d.day, t.hour, t.minute)
            due_at = due_dt.isoformat(timespec="seconds")
        elif isinstance(d, date) and t is None:
            due_dt = datetime(d.year, d.month, d.day, 23, 59)
            due_at = due_dt.isoformat(timespec="seconds")

    explanation = (st.session_state.get(f"new_explain_{sk}", "") or "").strip()

    if not title:
        set_teacher_notice(subject, "error", "Thiếu **tiêu đề quiz**.")
        return
    if not question:
        set_teacher_notice(subject, "error", "Thiếu **câu hỏi**.")
        return

    quiz = {
        "id": uuid.uuid4().hex,
        "subject": subject,
        "title": title,
        "question": question,
        "mandatory": mandatory,
        "difficulty": difficulty,
        "points": points,
        "max_attempts": max_attempts,  # 0 = unlimited
        "due_at": due_at,              # iso string or None
        "explanation": explanation,
        "created_at": datetime.now().isoformat(timespec="seconds"),
    }

    if qtype_ui == QUIZ_TYPE_MCQ:
        optA = (st.session_state.get(f"new_optA_{sk}", "") or "").strip()
        optB = (st.session_state.get(f"new_optB_{sk}", "") or "").strip()
        optC = (st.session_state.get(f"new_optC_{sk}", "") or "").strip()
        optD = (st.session_state.get(f"new_optD_{sk}", "") or "").strip()
        options = {"A": optA, "B": optB, "C": optC, "D": optD}

        non_empty = [k for k, v in options.items() if v]
        if len(non_empty) < 2:
            set_teacher_notice(subject, "error", "Quiz MCQ cần **ít nhất 2 lựa chọn** (A/B/C/D).")
            return

        correct_letter = st.session_state.get(f"new_correct_{sk}", "A")
        if correct_letter not in MCQ_LETTERS:
            correct_letter = "A"
        if not options.get(correct_letter):
            set_teacher_notice(subject, "error", f"Bạn chọn đáp án đúng **{correct_letter}** nhưng lựa chọn này đang trống.")
            return

        quiz["type"] = "mcq"
        quiz["options"] = options
        quiz["correct_letter"] = correct_letter

    else:
        answer_text = (st.session_state.get(f"new_answer_{sk}", "") or "").strip()
        if not answer_text:
            set_teacher_notice(subject, "error", "Quiz tự luận cần có **đáp án mẫu**.")
            return
        quiz["type"] = "short"
        quiz["answer_text"] = answer_text

    bank = st.session_state.quiz_bank
    bank.setdefault(subject, [])
    bank[subject] = bank[subject] + [quiz]
    st.session_state.quiz_bank = bank

    # clear inputs (callback-safe)
    st.session_state[f"new_title_{sk}"] = ""
    st.session_state[f"new_question_{sk}"] = ""
    st.session_state[f"new_explain_{sk}"] = ""
    st.session_state[f"new_mand_{sk}"] = False
    st.session_state[f"new_diff_{sk}"] = 3
    st.session_state[f"new_points_{sk}"] = 10
    st.session_state[f"new_unlim_{sk}"] = False
    st.session_state[f"new_attempts_{sk}"] = 1
    st.session_state[f"new_has_due_{sk}"] = False

    st.session_state[f"new_optA_{sk}"] = ""
    st.session_state[f"new_optB_{sk}"] = ""
    st.session_state[f"new_optC_{sk}"] = ""
    st.session_state[f"new_optD_{sk}"] = ""
    st.session_state[f"new_correct_{sk}"] = "A"
    st.session_state[f"new_answer_{sk}"] = ""

    set_teacher_notice(subject, "success", f"Đã tạo quiz: **{title}**")

def delete_quiz(subject: str, quiz_id: str):
    bank = st.session_state.quiz_bank
    old = bank.get(subject, [])
    bank[subject] = [q for q in old if q.get("id") != quiz_id]
    st.session_state.quiz_bank = bank

    subs = st.session_state.get("quiz_submissions", {})
    for sid in list(subs.keys()):
        subs[sid].pop(quiz_id, None)
    st.session_state.quiz_submissions = subs

    set_teacher_notice(subject, "success", "Đã xoá quiz.")

def toggle_quiz_mandatory(subject: str, quiz_id: str):
    bank = st.session_state.quiz_bank
    lst = list(bank.get(subject, []))
    new_lst = []
    changed = False
    for q in lst:
        if q.get("id") == quiz_id:
            q2 = dict(q)
            q2["mandatory"] = not bool(q2.get("mandatory"))
            new_lst.append(q2)
            changed = True
        else:
            new_lst.append(q)
    bank[subject] = new_lst
    st.session_state.quiz_bank = bank
    if changed:
        set_teacher_notice(subject, "info", "Đã đổi trạng thái bắt buộc của quiz.")


# =========================
# Submissions / grading (attempts + score + late)
# =========================
def set_student_notice(student_id: str, kind: str, msg: str):
    st.session_state.student_notice = {"student_id": student_id, "kind": kind, "msg": msg}

def get_submission(student_id: str, quiz_id: str):
    subs = st.session_state.get("quiz_submissions", {})
    return subs.get(student_id, {}).get(quiz_id)

def get_attempts(student_id: str, quiz_id: str):
    sub = get_submission(student_id, quiz_id)
    if not sub:
        return []
    return list(sub.get("attempts", []))

def best_score(student_id: str, quiz_id: str) -> int:
    atts = get_attempts(student_id, quiz_id)
    if not atts:
        return 0
    return int(max(a.get("score", 0) for a in atts))

def any_correct(student_id: str, quiz_id: str) -> bool:
    return any(a.get("correct") for a in get_attempts(student_id, quiz_id))

def last_attempt(student_id: str, quiz_id: str):
    atts = get_attempts(student_id, quiz_id)
    return atts[-1] if atts else None

def attempts_used(student_id: str, quiz_id: str) -> int:
    return len(get_attempts(student_id, quiz_id))

def attempts_left(student_id: str, quiz: dict) -> str:
    maxa = int(quiz.get("max_attempts", 1))
    used = attempts_used(student_id, quiz["id"])
    if maxa <= 0:
        return "∞"
    left = max(0, maxa - used)
    return str(left)

def grade_quiz(student_id: str, subject: str, quiz_id: str):
    q = find_quiz(subject, quiz_id)
    if not q:
        set_student_notice(student_id, "error", "Không tìm thấy quiz (có thể đã bị xoá).")
        return

    maxa = int(q.get("max_attempts", 1))
    used = attempts_used(student_id, quiz_id)
    if maxa > 0 and used >= maxa:
        set_student_notice(student_id, "error", f"Bạn đã hết lượt làm quiz này. (Max attempts = {maxa})")
        return

    due_dt = parse_due(q.get("due_at"))
    late = bool(due_dt and datetime.now() > due_dt)

    ans_key = f"ans_{student_id}_{quiz_id}"
    user_ans = st.session_state.get(ans_key, None)

    correct = False
    correct_answer_show = ""

    if q.get("type") == "mcq":
        options = q.get("options", {}) or {}
        if user_ans is None:
            set_student_notice(student_id, "error", "Bạn chưa chọn đáp án.")
            return
        correct_letter = q.get("correct_letter", "A")
        correct = (user_ans == correct_letter)
        correct_answer_show = f"{correct_letter}. {options.get(correct_letter,'')}"
    else:
        user_ans = (user_ans or "").strip()
        if not user_ans:
            set_student_notice(student_id, "error", "Bạn chưa nhập câu trả lời.")
            return
        correct_text = (q.get("answer_text") or "").strip()
        correct = (normalize_text(user_ans) == normalize_text(correct_text))
        correct_answer_show = correct_text

    pts = int(q.get("points", 10))
    score = pts if correct else 0

    subs = st.session_state.quiz_submissions
    subs.setdefault(student_id, {})
    subs.setdefault(student_id, {}).setdefault(quiz_id, {"subject": subject, "attempts": []})

    subs[student_id][quiz_id]["subject"] = subject
    subs[student_id][quiz_id]["attempts"] = subs[student_id][quiz_id].get("attempts", []) + [{
        "answer": user_ans,
        "correct": bool(correct),
        "score": int(score),
        "late": bool(late),
        "submitted_at": datetime.now().isoformat(timespec="seconds"),
    }]
    st.session_state.quiz_submissions = subs

    left = attempts_left(student_id, q)
    late_note = " (NỘP MUỘN)" if late else ""
    if correct:
        set_student_notice(student_id, "success", f"✅ Đúng rồi! +{score}đ{late_note} | Còn lượt: {left}")
    else:
        msg = f"❌ Chưa đúng. +0đ{late_note}\nĐáp án đúng: {correct_answer_show}\nCòn lượt: {left}"
        if q.get("explanation"):
            msg += f"\n\nGiải thích: {q.get('explanation')}"
        set_student_notice(student_id, "warning", msg)

def get_pending_mandatory_quizzes(student_id: str):
    pending = []
    for subj in ALL_SUBJECTS:
        for q in get_quizzes(subj):
            if not q.get("mandatory"):
                continue
            if attempts_used(student_id, q["id"]) == 0:
                pending.append((subj, q))
    pending.sort(key=lambda x: (ALL_SUBJECTS.index(x[0]), safe_str(x[1].get("due_at") or "")))
    return pending


# =========================
# Study items for schedule (quiz ưu tiên trước)
# =========================
def get_study_items_for_subject(subject: str):
    quizzes = get_quizzes(subject)
    quizzes_sorted = sorted(
        quizzes,
        key=lambda q: (not bool(q.get("mandatory")), safe_str(q.get("due_at") or ""), safe_str(q.get("created_at") or "")),
    )
    quiz_tasks = [quiz_task_title(q) for q in quizzes_sorted]

    defaults = EXERCISES_DEFAULT.get(subject, [])
    seen = set()
    merged = []
    for x in quiz_tasks + defaults:
        if x not in seen:
            merged.append(x)
            seen.add(x)
    return merged

def pick_study_items(subject: str, n: int = 2):
    items = get_study_items_for_subject(subject)
    return items[:n] if len(items) >= n else items


# =========================
# Smart schedule generator (T2–CN)
# =========================
def allocate_weekly_session_counts(
    scores: dict,
    total_slots: int = 14,
    min_count: int = 1,
    max_count: int = 4,
    base_priority: float = 0.5
):
    subjects = ALL_SUBJECTS[:]
    counts = {s: min_count for s in subjects}
    extras = total_slots - min_count * len(subjects)
    if extras <= 0:
        return counts

    priority = {}
    for s in subjects:
        score = float(scores.get(s, 0))
        deficit = max(0.0, PASS_THRESHOLD - score)
        priority[s] = base_priority + deficit

    for _ in range(extras):
        candidates = [s for s in subjects if counts[s] < max_count]
        if not candidates:
            candidates = subjects

        def key(s):
            score = float(scores.get(s, 0))
            ratio = priority[s] / max(counts[s], 1)
            return (ratio, -score, -subjects.index(s))

        pick = max(candidates, key=key)
        counts[pick] += 1

    return counts

def build_weekly_schedule_subjects(scores: dict, days=None, sessions_per_day: int = 2):
    if days is None:
        days = ["T2", "T3", "T4", "T5", "T6", "T7", "CN"]

    total_slots = len(days) * sessions_per_day
    counts = allocate_weekly_session_counts(scores, total_slots=total_slots)

    remaining = dict(counts)
    schedule = []
    prev_day = set()

    for day in days:
        today = []
        for _ in range(sessions_per_day):
            candidates = [s for s in ALL_SUBJECTS if remaining.get(s, 0) > 0 and s not in today]
            if not candidates:
                candidates = [s for s in ALL_SUBJECTS if remaining.get(s, 0) > 0]

            cand2 = [s for s in candidates if s not in prev_day]
            if cand2:
                candidates = cand2

            def key(s):
                score = float(scores.get(s, 0))
                return (remaining.get(s, 0), -score, -ALL_SUBJECTS.index(s))

            pick = max(candidates, key=key)
            today.append(pick)
            remaining[pick] -= 1

        schedule.append((day, today))
        prev_day = set(today)

    return schedule

def attach_tasks_to_schedule(schedule):
    idx = {s: 0 for s in ALL_SUBJECTS}
    out = []
    for day, subs in schedule:
        sessions = []
        for s in subs:
            items = get_study_items_for_subject(s)
            task = items[idx[s] % len(items)] if items else "Ôn lý thuyết + làm bài trong SGK/vở bài tập."
            if items:
                idx[s] += 1
            sessions.append((s, task))
        out.append((day, sessions))
    return out


# =========================
# AI chat (rule-based)
# =========================
def hk2_revision_plan(input_row: dict, n_each: int = 3) -> str:
    scores = {s: float(input_row.get(s, 0)) for s in ALL_SUBJECTS}
    low = get_low_subjects(scores, threshold=PASS_THRESHOLD)

    out = []
    out.append("📚 ÔN TẬP HK2 - GỢI Ý + LỊCH ÔN (T2–CN)")
    out.append("")

    if low:
        focus = low[:min(len(low), 3)]
        out.append(
            f"Ưu tiên vì đang dưới {PASS_THRESHOLD}: "
            + ", ".join([f"{DISPLAY_NAME.get(s, s)}({v})" for s, v in focus])
            + ("..." if len(low) > 3 else "")
        )
    else:
        out.append(f"Tất cả môn đang ≥ {PASS_THRESHOLD}. Lịch sẽ chia đều hơn.")

    out.append("")
    out.append("1) GỢI Ý THEO MÔN (mỗi môn 3 mục) — ưu tiên quiz bắt buộc nếu có:")
    out.append("")

    for subj in ALL_SUBJECTS:
        disp = DISPLAY_NAME.get(subj, subj)
        out.append(f"{disp}:")
        items = get_study_items_for_subject(subj)
        for item in items[:n_each]:
            out.append(f"• {item}")
        out.append("")

    out.append("2) LỊCH ÔN THEO NGÀY (T2–CN):")
    out.append("")

    schedule = build_weekly_schedule_subjects(scores)
    schedule_with_tasks = attach_tasks_to_schedule(schedule)
    for day, sessions in schedule_with_tasks:
        out.append(f"{day}:")
        for subj, task in sessions:
            out.append(f"• {DISPLAY_NAME.get(subj, subj)}: {task}")
        out.append("")

    return "\n".join(out)

def full_recommendation(current, input_row) -> str:
    logic_avg, lang_avg, diff, level = compute_scores({**current, **input_row})
    scores_for_pick = {s: float(input_row.get(s, 0)) for s in ALL_SUBJECTS}
    low = get_low_subjects(scores_for_pick, threshold=PASS_THRESHOLD)

    out = []
    out.append(f"Kết quả: TB Logic={logic_avg} | TB Ngôn ngữ={lang_avg} | Δ={diff}")
    out.append(f"Cân bằng nhóm: {level}")
    out.append("")

    if low:
        out.append(f"🛑 CẢNH BÁO: Có {len(low)} môn dưới {PASS_THRESHOLD} → cần ưu tiên cải thiện.")
        focus = low[:min(FOCUS_K, len(low))]
        out.append("Môn ưu tiên (thấp nhất trước):")
        out.append("• " + ", ".join([f"{DISPLAY_NAME.get(s, s)} ({v})" for s, v in focus]))
        out.append("")
        out.append("Gợi ý (mỗi môn 2 mục) — ưu tiên quiz bắt buộc nếu có:")
        for subj, _v in focus:
            disp = DISPLAY_NAME.get(subj, subj)
            out.append(f"{disp}:")
            for item in pick_study_items(subj, n=2):
                out.append(f"• {item}")
        out.append("")
        out.append("Gợi ý: Gõ “Tôi cần ôn các bài tập gì?” để lấy lịch ôn T2–CN.")
        return "\n".join(out)

    if level == "Cân bằng tốt":
        out += [
            "✅ Gợi ý (2–3 tuần):",
            "• Duy trì nhịp học hiện tại.",
            "• Tăng thử thách nhẹ ở môn mạnh.",
            "• Theo dõi lại sau 2 tuần.",
        ]
        return "\n".join(out)

    if level == "Lệch vừa":
        out += [
            "⚠️ Gợi ý (2–3 tuần):",
            "• Tăng 15–20 phút/ngày cho nhóm yếu.",
            "• Theo dõi lại sau 2 tuần.",
        ]
        return "\n".join(out)

    weak_group = "Ngôn ngữ" if logic_avg > lang_avg else "Logic"
    out += [
        "🛑 Gợi ý ưu tiên (Lệch rõ):",
        f"• Nhóm yếu hiện tại: {weak_group}",
        "• Trong 2 tuần: ưu tiên kéo Δ xuống mức 'Lệch vừa'.",
        "",
    ]
    weak2 = weakest_subjects(scores_for_pick, "Logic" if weak_group == "Logic" else "Ngôn ngữ", k=2)
    out.append("2 môn yếu nhất & gợi ý (ưu tiên quiz bắt buộc):")
    for subj in weak2:
        disp = DISPLAY_NAME.get(subj, subj)
        out.append(f"{disp}:")
        for item in pick_study_items(subj, n=2):
            out.append(f"• {item}")
    out.append("")
    out.append("Gợi ý: Gõ “Tôi cần ôn các bài tập gì?” để lấy lịch ôn T2–CN.")
    return "\n".join(out)

def detect_subject_from_text(tn: str):
    patterns = [
        ("Toán", ["mon toan", "toan hoc", "diem toan", "math"]),
        ("Khoa học", ["mon khoa hoc", "khoa hoc", "science", "sci"]),
        ("Tin", ["mon tin", "tin hoc", "lap trinh", "programming", "program", "code", "it", "computer"]),
        ("Văn", ["mon van", "ngu van", "literature", "doc hieu", "viet van"]),
        ("Sử", ["mon su", "lich su", "history"]),
        ("Anh", ["mon anh", "tieng anh", "english"]),
    ]
    for subj, pats in patterns:
        for p in pats:
            if p in tn:
                return subj
    return None

def quiz_pack_for_subject(subj: str, max_quiz: int = 5) -> str:
    qs = get_quizzes(subj)
    if not qs:
        return f"Hiện chưa có quiz nào cho **{DISPLAY_NAME.get(subj, subj)}**."

    qs_sorted = sorted(qs, key=lambda q: (not bool(q.get("mandatory")), safe_str(q.get("due_at") or ""), safe_str(q.get("created_at") or "")))
    out = []
    out.append(f"🧩 Quiz cho {DISPLAY_NAME.get(subj, subj)} (tối đa {max_quiz})")
    out.append("(Quiz bắt buộc hiển thị trước)")
    out.append("")
    for i, q in enumerate(qs_sorted[:max_quiz], start=1):
        diff = int(q.get("difficulty", 3))
        pts = int(q.get("points", 10))
        due = format_due(parse_due(q.get("due_at")))
        out.append(f"{i}) {quiz_badge(q)} — ⭐{diff}/5 — {pts}đ — {due} — **{q.get('title','(No title)')}**")
        out.append(f"   Câu hỏi: {q.get('question','')}")
        if q.get("type") == "mcq":
            opts = q.get("options", {}) or {}
            for L in MCQ_LETTERS:
                if opts.get(L):
                    out.append(f"   {L}. {opts[L]}")
        out.append("")
    return "\n".join(out)

def quiz_pack_all(max_each: int = 2) -> str:
    out = []
    out.append("🧩 Tổng hợp quiz theo môn")
    out.append("")
    for subj in ALL_SUBJECTS:
        qs = get_quizzes(subj)
        disp = DISPLAY_NAME.get(subj, subj)
        out.append(f"{disp}:")
        if not qs:
            out.append("• (Chưa có quiz)")
        else:
            qs_sorted = sorted(qs, key=lambda q: (not bool(q.get("mandatory")), safe_str(q.get("due_at") or ""), safe_str(q.get("created_at") or "")))
            for q in qs_sorted[:max_each]:
                diff = int(q.get("difficulty", 3))
                pts = int(q.get("points", 10))
                out.append(f"• {quiz_badge(q)} — ⭐{diff}/5 — {pts}đ — {q.get('title','(No title)')}")
        out.append("")
    return "\n".join(out)

def ai_chat_answer(user_text: str, current, input_row) -> str:
    tn = normalize_text(user_text)

    if tn in ["hi", "hello", "hey", "xin chao", "chao", "alo", "yo", "hiii", "helo"]:
        return (
            "Chào bạn 👋\n\n"
            "Bạn thử hỏi:\n"
            "• Mình đang yếu môn nào?\n"
            "• Cho recommendation giúp mình\n"
            "• Tôi cần ôn các bài tập gì?\n"
            "• Cho mình quiz môn Toán / Tiếng Anh / ...\n"
        )

    if (
        "toi can on cac bai tap gi" in tn
        or "toi can on bai tap gi" in tn
        or ("on" in tn and "bai tap" in tn and ("hoc ky 2" in tn or "hoc ki 2" in tn or "hk2" in tn))
    ):
        return hk2_revision_plan(input_row, n_each=3)

    if "quiz" in tn or "cau hoi" in tn:
        subj = detect_subject_from_text(tn)
        if subj:
            return quiz_pack_for_subject(subj, max_quiz=5)
        return quiz_pack_all(max_each=2)

    trigger = any(k in tn for k in [
        "yeu", "mon nao", "mon yeu", "yeu mon", "diem thap", "lech", "can bang",
        "goi y", "recommend", "recommendation", "tom tat", "tong hop", "ke hoach",
        "improve", "cai thien", "duoi", "thap"
    ])
    if trigger:
        return full_recommendation(current, input_row)

    return (
        "Bạn hỏi thử kiểu:\n"
        "• Mình đang yếu môn nào?\n"
        "• Cho recommendation giúp mình\n"
        "• Tôi cần ôn các bài tập gì?\n"
        "• Cho mình quiz môn Toán / Tiếng Anh / ...\n"
    )


# =========================
# Teacher dashboard stats
# =========================
def teacher_stats_rows():
    quizzes = all_quizzes()
    students = st.session_state.get("students", [])
    student_ids = [s["ID"] for s in students]

    mandatory_quizzes = [q for q in quizzes if q.get("mandatory")]
    mandatory_total = len(mandatory_quizzes)

    # per-student rows
    per_student = []
    for s in students:
        sid = s["ID"]
        submitted_any = 0
        total_best_score = 0

        mand_done = 0
        mand_pending = 0
        overdue_pending = 0

        for q in quizzes:
            qid = q["id"]
            if attempts_used(sid, qid) > 0:
                submitted_any += 1
                total_best_score += best_score(sid, qid)
            if q.get("mandatory"):
                if attempts_used(sid, qid) > 0:
                    mand_done += 1
                else:
                    mand_pending += 1
                    due_dt = parse_due(q.get("due_at"))
                    if due_dt and datetime.now() > due_dt:
                        overdue_pending += 1

        completion = (mand_done / mandatory_total * 100.0) if mandatory_total > 0 else 0.0
        avg_score_attempted = (total_best_score / submitted_any) if submitted_any > 0 else 0.0

        per_student.append({
            "ID": sid,
            "Lớp": s.get("Lớp", ""),
            "Tên": s.get("Tên", ""),
            "Mandatory done": mand_done,
            "Mandatory pending": mand_pending,
            "Overdue pending": overdue_pending,
            "Completion %": round(completion, 1),
            "Submitted quizzes": submitted_any,
            "Total best score": total_best_score,
            "Avg score/attempted": round(avg_score_attempted, 2),
        })

    # per-quiz rows
    per_quiz = []
    for q in quizzes:
        qid = q["id"]
        attempted = 0
        correct_any_cnt = 0
        best_scores = []
        attempts_counts = []
        late_cnt = 0
        due_dt = parse_due(q.get("due_at"))
        overdue_pending = 0

        for sid in student_ids:
            atts = get_attempts(sid, qid)
            if atts:
                attempted += 1
                attempts_counts.append(len(atts))
                bs = best_score(sid, qid)
                best_scores.append(bs)
                if any(a.get("correct") for a in atts):
                    correct_any_cnt += 1
                if any(a.get("late") for a in atts):
                    late_cnt += 1
            else:
                if q.get("mandatory") and due_dt and datetime.now() > due_dt:
                    overdue_pending += 1

        correct_rate = (correct_any_cnt / attempted * 100.0) if attempted > 0 else 0.0
        avg_best = (sum(best_scores) / len(best_scores)) if best_scores else 0.0
        avg_attempts = (sum(attempts_counts) / len(attempts_counts)) if attempts_counts else 0.0

        per_quiz.append({
            "Subject": q.get("subject", ""),
            "Title": q.get("title", ""),
            "Mandatory": bool(q.get("mandatory")),
            "Difficulty": int(q.get("difficulty", 3)),
            "Points": int(q.get("points", 10)),
            "MaxAttempts": int(q.get("max_attempts", 1)),
            "Due": format_due(due_dt),
            "Students attempted": attempted,
            "Correct rate %": round(correct_rate, 1),
            "Avg best score": round(avg_best, 2),
            "Avg attempts": round(avg_attempts, 2),
            "Late submitters": late_cnt,
            "Overdue mandatory pending": overdue_pending,
        })

    return per_student, per_quiz


# =========================
# Session init
# =========================
if "students" not in st.session_state:
    st.session_state.students = [
        {"ID": "A01", "Lớp": "8A", "Tên": "Học sinh A",
         "Toán": 9.0, "Khoa học": 8.0, "Tin": 7.5, "Văn": 5.5, "Sử": 7.0, "Anh": 6.0},
        {"ID": "B02", "Lớp": "8A", "Tên": "Học sinh B",
         "Toán": 7.5, "Khoa học": 7.2, "Tin": 7.0, "Văn": 7.0, "Sử": 6.8, "Anh": 7.1},
        {"ID": "C03", "Lớp": "8B", "Tên": "Học sinh C",
         "Toán": 9.5, "Khoa học": 9.0, "Tin": 9.0, "Văn": 5.0, "Sử": 5.5, "Anh": 5.0},
    ]

if "quiz_bank" not in st.session_state:
    st.session_state.quiz_bank = {s: [] for s in ALL_SUBJECTS}

if "quiz_submissions" not in st.session_state:
    # student_id -> quiz_id -> {"subject":..., "attempts":[...]}
    st.session_state.quiz_submissions = {}

if "teacher_notice" not in st.session_state:
    st.session_state.teacher_notice = None
if "import_notice" not in st.session_state:
    st.session_state.import_notice = None
if "student_notice" not in st.session_state:
    st.session_state.student_notice = None

for subj in ALL_SUBJECTS:
    sk = SUBJ_KEY[subj]
    st.session_state.setdefault(f"new_title_{sk}", "")
    st.session_state.setdefault(f"new_question_{sk}", "")
    st.session_state.setdefault(f"new_type_{sk}", QUIZ_TYPE_MCQ)
    st.session_state.setdefault(f"new_mand_{sk}", False)
    st.session_state.setdefault(f"new_diff_{sk}", 3)
    st.session_state.setdefault(f"new_points_{sk}", 10)
    st.session_state.setdefault(f"new_unlim_{sk}", False)
    st.session_state.setdefault(f"new_attempts_{sk}", 1)
    st.session_state.setdefault(f"new_has_due_{sk}", False)
    st.session_state.setdefault(f"new_due_date_{sk}", date.today())
    st.session_state.setdefault(f"new_due_time_{sk}", dt_time(23, 59))
    st.session_state.setdefault(f"new_explain_{sk}", "")

    st.session_state.setdefault(f"new_optA_{sk}", "")
    st.session_state.setdefault(f"new_optB_{sk}", "")
    st.session_state.setdefault(f"new_optC_{sk}", "")
    st.session_state.setdefault(f"new_optD_{sk}", "")
    st.session_state.setdefault(f"new_correct_{sk}", "A")
    st.session_state.setdefault(f"new_answer_{sk}", "")

if "chat_history" not in st.session_state:
    st.session_state.chat_history = [
        {"role": "assistant", "content": "Xin chào! Bạn có thể hỏi: “Cho mình quiz môn Toán” hoặc “Tôi cần ôn các bài tập gì?”."}
    ]
if "ai_question" not in st.session_state:
    st.session_state.ai_question = ""
if "current_student" not in st.session_state:
    st.session_state.current_student = None
if "current_input_row" not in st.session_state:
    st.session_state.current_input_row = None
if "chat_minimized" not in st.session_state:
    st.session_state.chat_minimized = False
if "chat_autoscroll" not in st.session_state:
    st.session_state.chat_autoscroll = True
if "chat_body_id" not in st.session_state:
    st.session_state.chat_body_id = f"chat-body-{uuid.uuid4().hex}"


def send_ai_message():
    q = st.session_state.get("ai_question", "").strip()
    if not q:
        return
    cur = st.session_state.get("current_student")
    row = st.session_state.get("current_input_row")
    if cur is None or row is None:
        return
    st.session_state.chat_history.append({"role": "user", "content": q})
    st.session_state.chat_history.append({"role": "assistant", "content": ai_chat_answer(q, cur, row)})
    st.session_state.ai_question = ""
    st.session_state.chat_autoscroll = True
    st.session_state.chat_minimized = False

def clear_ai_chat():
    st.session_state.chat_history = [
        {"role": "assistant", "content": "Chat đã xoá. Bạn có thể hỏi: “Cho mình quiz môn Toán” hoặc “Tôi cần ôn các bài tập gì?”."}
    ]
    st.session_state.ai_question = ""
    st.session_state.chat_autoscroll = True

def toggle_minimize():
    st.session_state.chat_minimized = not st.session_state.chat_minimized


# =========================
# Main layout
# =========================
left_col, mid_col, right_col = st.columns([0.70, 3.10, 1.40])

# LEFT NAV
with left_col:
    st.markdown(
        """
        <div class="leftNav">
          <div class="brand">VINSCHOOL</div>
          <div class="navItem">👤 Tài Khoản<br><small>Account</small></div>
          <div class="navItem muted">📊 Bảng Điều Khiển<br><small>Dashboard</small></div>
          <div class="navItem">📚 Khóa Học<br><small>Courses</small></div>
          <div class="navItem">👥 Các nhóm<br><small>Groups</small></div>
          <div class="navItem">📅 Lịch<br><small>Calendar</small></div>
          <div class="navItem">📥 Hộp thư đến<br><small>Inbox</small></div>
          <div class="navItem">🕘 Lịch sử<br><small>History</small></div>
          <div style="margin-top:14px; opacity:0.75; font-size:12px;">Replica UI để demo</div>
        </div>
        """,
        unsafe_allow_html=True
    )

# MIDDLE
with mid_col:
    st.markdown("# Bảng Điều Khiển")

    # Cards
    r1 = st.columns(3)
    r2 = st.columns(3)
    for i, subj in enumerate(SUBJECT_CARDS):
        target = r1[i] if i < 3 else r2[i - 3]
        with target:
            st.markdown(
                f"""
                <div class="course-card">
                  <div class="course-top" style="background:{subj["color"]};">
                    <div class="course-dots">⋮</div>
                  </div>
                  <div class="course-body">
                    <div class="course-slug">{subj["slug"]}</div>
                    <div class="course-code">{subj["name"]}<br>{subj["code"]}<br>{subj["year"]}</div>
                    <div class="course-icons"><span>📝</span><span>💬</span><span>📁</span></div>
                  </div>
                </div>
                """,
                unsafe_allow_html=True
            )

    # Balance tool
    st.markdown('<div class="midPanel">', unsafe_allow_html=True)
    st.markdown("## Thước Cân Bằng Học Tập")
    st.caption(
        f"Rule: Nếu có bất kỳ môn nào < {PASS_THRESHOLD} → báo cần cải thiện + gợi ý. "
        f"(Quiz giáo viên tạo nằm ở phần Teacher tools bên dưới.)"
    )

    labels = [student_label(s) for s in st.session_state.students]
    selected = st.selectbox("Chọn học sinh", labels, key="selected_student")
    current = next(s for s in st.session_state.students if student_label(s) == selected)

    if st.session_state.get("loaded_student_id") != current["ID"]:
        st.session_state.loaded_student_id = current["ID"]
        st.session_state.score_math = float(current["Toán"])
        st.session_state.score_science = float(current["Khoa học"])
        st.session_state.score_it = float(current["Tin"])
        st.session_state.score_lit = float(current["Văn"])
        st.session_state.score_history = float(current["Sử"])
        st.session_state.score_english = float(current["Anh"])

    colA, colB = st.columns(2)
    with colA:
        st.markdown("**Logic – STEM**")
        math = st.number_input("Toán", 0.0, 10.0, step=0.1, key="score_math")
        sci = st.number_input("Khoa học", 0.0, 10.0, step=0.1, key="score_science")
        it = st.number_input("Tin", 0.0, 10.0, step=0.1, key="score_it")
    with colB:
        st.markdown("**Ngôn ngữ – Xã hội**")
        lit = st.number_input("Văn", 0.0, 10.0, step=0.1, key="score_lit")
        hist = st.number_input("Sử", 0.0, 10.0, step=0.1, key="score_history")
        eng = st.number_input("Anh", 0.0, 10.0, step=0.1, key="score_english")

    input_row = {"Toán": math, "Khoa học": sci, "Tin": it, "Văn": lit, "Sử": hist, "Anh": eng}
    st.session_state.current_student = current
    st.session_state.current_input_row = input_row

    logic_avg, lang_avg, diff, level = compute_scores({**current, **input_row})
    low_ui = get_low_subjects(input_row, threshold=PASS_THRESHOLD)

    k1, k2, k3 = st.columns(3)
    k1.metric("TB Logic", logic_avg)
    k2.metric("TB Ngôn ngữ", lang_avg)
    k3.metric("Δ", diff)

    if low_ui:
        preview = ", ".join([f"{DISPLAY_NAME.get(s, s)}({v})" for s, v in low_ui[:3]]) + ("..." if len(low_ui) > 3 else "")
        st.error(f"Cần cải thiện: Có {len(low_ui)} môn dưới {PASS_THRESHOLD}: {preview}")
    else:
        st.success(level) if level == "Cân bằng tốt" else st.warning(level) if level == "Lệch vừa" else st.error(level)

    st.progress(min(diff / 3, 1.0))
    st.caption("Ngưỡng cân bằng theo Δ: Δ < 1 (tốt), 1 ≤ Δ < 2 (vừa), Δ ≥ 2 (rõ)")

    fig, ax = plt.subplots(figsize=(6, 3.2))
    ax.bar(["Logic", "Ngôn ngữ"], [logic_avg, lang_avg])
    ax.set_ylim(0, 10)
    ax.set_ylabel("Điểm TB")
    st.pyplot(fig)

    st.markdown("</div>", unsafe_allow_html=True)

    # =========================
    # Teacher tools
    # =========================
    st.markdown('<div class="midPanel">', unsafe_allow_html=True)
    st.markdown("## 👩‍🏫 Teacher tools: Quiz Bank + Deadline + Points + Attempts + Excel")

    st.caption(
        "• MCQ: A/B/C/D + đáp án đúng\n"
        "• Difficulty 1–5 ⭐\n"
        "• Points (điểm)\n"
        "• Deadline (tùy chọn)\n"
        "• Max attempts (0 = Unlimited)\n"
        "• Import/Export Excel"
    )

    # Excel import/export
    st.markdown("### 📥📤 Excel Import / Export")

    cex1, cex2 = st.columns([1, 1])

    with cex1:
        tpl = make_template_xlsx_bytes()
        if tpl:
            st.download_button(
                "⬇️ Tải mẫu Excel",
                data=tpl,
                file_name="quiz_template.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
        else:
            st.info("Không tạo được template (thiếu openpyxl).")

    with cex2:
        exp = export_quiz_bank_xlsx_bytes()
        if exp:
            st.download_button(
                "⬇️ Export quiz bank ra Excel",
                data=exp,
                file_name="quiz_bank_export.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
        else:
            st.info("Không export được (thiếu openpyxl).")

    st.file_uploader("Upload file Excel để import", type=["xlsx"], key="import_excel_file")
    st.button("📥 Import ngay", on_click=import_quizzes_from_excel, use_container_width=True)

    imp = st.session_state.get("import_notice")
    if imp:
        kind = imp.get("kind", "info")
        msg = imp.get("msg", "")
        details = imp.get("details")
        if kind == "success":
            st.success(msg)
        elif kind == "warning":
            st.warning(msg)
        elif kind == "error":
            st.error(msg)
        else:
            st.info(msg)
        if details:
            with st.expander("Xem chi tiết lỗi"):
                st.code(details)

    st.divider()

    # Teacher Dashboard
    st.markdown("### 📊 Teacher Dashboard")
    per_student, per_quiz = teacher_stats_rows()

    m1, m2, m3, m4 = st.columns(4)
    total_quiz = len(all_quizzes())
    mand_total = sum(1 for q in all_quizzes() if q.get("mandatory"))
    m1.metric("Total quizzes", total_quiz)
    m2.metric("Mandatory quizzes", mand_total)
    m3.metric("Total students", len(st.session_state.students))
    # overall completion avg
    avg_completion = 0.0
    if per_student:
        avg_completion = sum(r["Completion %"] for r in per_student) / len(per_student)
    m4.metric("Avg mandatory completion", f"{avg_completion:.1f}%")

    st.markdown("**Theo học sinh**")
    st.dataframe(per_student, use_container_width=True, hide_index=True)

    st.markdown("**Theo quiz**")
    st.dataframe(per_quiz, use_container_width=True, hide_index=True)

    st.divider()

    # Per-subject tabs for creating quizzes
    tabs = st.tabs([DISPLAY_NAME.get(s, s) for s in ALL_SUBJECTS])
    for subj, tab in zip(ALL_SUBJECTS, tabs):
        sk = SUBJ_KEY[subj]
        disp = DISPLAY_NAME.get(subj, subj)

        with tab:
            notice = st.session_state.get("teacher_notice")
            if notice and notice.get("subject") == subj:
                kind = notice.get("kind", "info")
                msg = notice.get("msg", "")
                if kind == "success":
                    st.success(msg)
                elif kind == "error":
                    st.error(msg)
                elif kind == "warning":
                    st.warning(msg)
                else:
                    st.info(msg)

            st.markdown(f"### ➕ Tạo quiz mới cho {disp}")

            left, right = st.columns([2.2, 1])

            with left:
                st.text_input("Tiêu đề quiz", key=f"new_title_{sk}")
                st.text_area("Câu hỏi", key=f"new_question_{sk}", height=120)

                st.selectbox("Loại quiz", QUIZ_TYPES, key=f"new_type_{sk}")

                a1, a2, a3 = st.columns(3)
                with a1:
                    st.slider("Độ khó ⭐ (1–5)", 1, 5, key=f"new_diff_{sk}")
                with a2:
                    st.number_input("Points (điểm)", min_value=1, max_value=1000, step=1, key=f"new_points_{sk}")
                with a3:
                    st.checkbox("Unlimited attempts", key=f"new_unlim_{sk}")
                    st.number_input("Max attempts", min_value=1, max_value=99, step=1, key=f"new_attempts_{sk}", disabled=bool(st.session_state.get(f"new_unlim_{sk}", False)))

                st.checkbox("Có deadline?", key=f"new_has_due_{sk}")
                if st.session_state.get(f"new_has_due_{sk}", False):
                    dcol, tcol = st.columns(2)
                    with dcol:
                        st.date_input("Due date", key=f"new_due_date_{sk}")
                    with tcol:
                        st.time_input("Due time", key=f"new_due_time_{sk}")

                qtype = st.session_state.get(f"new_type_{sk}", QUIZ_TYPE_MCQ)
                if qtype == QUIZ_TYPE_MCQ:
                    st.markdown("**Lựa chọn (A–D):**")
                    o1, o2 = st.columns(2)
                    with o1:
                        st.text_input("A", key=f"new_optA_{sk}")
                        st.text_input("C", key=f"new_optC_{sk}")
                    with o2:
                        st.text_input("B", key=f"new_optB_{sk}")
                        st.text_input("D", key=f"new_optD_{sk}")

                    st.selectbox("Đáp án đúng", MCQ_LETTERS, key=f"new_correct_{sk}")
                else:
                    st.text_area("Đáp án mẫu (để chấm/đối chiếu)", key=f"new_answer_{sk}", height=120)

                st.text_area("Giải thích (tuỳ chọn)", key=f"new_explain_{sk}", height=90)

            with right:
                st.markdown("### ✅ Trạng thái")
                st.checkbox("Quiz này bắt buộc làm?", key=f"new_mand_{sk}")
                st.button(
                    "➕ Tạo quiz",
                    key=f"btn_add_{sk}",
                    on_click=add_quiz_from_inputs,
                    args=(subj,),
                    use_container_width=True,
                )

                st.markdown("---")
                st.markdown("### 📌 Thống kê")
                st.write(f"Số quiz: **{len(get_quizzes(subj))}**")
                st.write(f"Bắt buộc: **{sum(1 for q in get_quizzes(subj) if q.get('mandatory'))}**")

            st.markdown("---")
            st.markdown(f"### 📚 Danh sách quiz của {disp}")

            quizzes = get_quizzes(subj)
            if not quizzes:
                st.info("Chưa có quiz nào. Tạo quiz hoặc import Excel.")
            else:
                quizzes_sorted = sorted(quizzes, key=lambda q: (not bool(q.get("mandatory")), safe_str(q.get("due_at") or ""), safe_str(q.get("created_at") or "")))
                for q in quizzes_sorted:
                    qid = q["id"]
                    badge = quiz_badge(q)
                    title = q.get("title", "(No title)")
                    qtype = q.get("type", "mcq")
                    diffq = int(q.get("difficulty", 3))
                    pts = int(q.get("points", 10))
                    maxa = int(q.get("max_attempts", 1))
                    due_dt = parse_due(q.get("due_at"))

                    due_txt = format_due(due_dt)
                    maxa_txt = "∞" if maxa <= 0 else str(maxa)

                    with st.expander(f"{badge} — ⭐{diffq}/5 — {pts}đ — attempts:{maxa_txt} — {due_txt} — {title}"):
                        st.markdown(f"**Câu hỏi:** {q.get('question','')}")
                        st.markdown(f"**Difficulty:** ⭐{diffq}/5")
                        st.markdown(f"**Points:** {pts} điểm")
                        st.markdown(f"**Max attempts:** {maxa_txt}")
                        st.markdown(f"**Deadline:** {due_txt}")

                        if qtype == "mcq":
                            opts = q.get("options", {}) or {}
                            st.markdown("**Lựa chọn:**")
                            for L in MCQ_LETTERS:
                                if opts.get(L):
                                    st.write(f"{L}. {opts[L]}")
                            cl = q.get("correct_letter", "A")
                            st.markdown(f"**Đáp án đúng:** {cl}. {opts.get(cl,'')}")
                        else:
                            st.markdown(f"**Đáp án mẫu:** {q.get('answer_text','')}")

                        if q.get("explanation"):
                            st.markdown(f"**Giải thích:** {q.get('explanation')}")

                        c1, c2 = st.columns([1, 1])
                        with c1:
                            st.button(
                                f"🔁 Bắt buộc? (Hiện tại: {badge})",
                                key=f"tog_{qid}",
                                on_click=toggle_quiz_mandatory,
                                args=(subj, qid),
                                use_container_width=True,
                            )
                        with c2:
                            st.button(
                                "🗑️ Xoá quiz",
                                key=f"del_{qid}",
                                on_click=delete_quiz,
                                args=(subj, qid),
                                use_container_width=True,
                            )

    st.markdown("</div>", unsafe_allow_html=True)

    # =========================
    # Student quiz demo
    # =========================
    st.markdown('<div class="midPanel">', unsafe_allow_html=True)
    st.markdown("## 🧑‍🎓 Quiz cho học sinh (demo/preview)")
    st.caption("Quiz bắt buộc ưu tiên hiển thị trước. Có deadline, points, attempts.")

    student_id = current["ID"]
    st.session_state.quiz_submissions.setdefault(student_id, {})

    f1, f2 = st.columns([1.2, 1])
    with f1:
        subj_filter = st.selectbox(
            "Lọc theo môn",
            ["Tất cả"] + [DISPLAY_NAME.get(s, s) for s in ALL_SUBJECTS],
            key="quiz_filter_subject",
        )
    with f2:
        only_mand = st.checkbox("Chỉ hiển thị quiz bắt buộc", key="quiz_filter_mand", value=False)

    s_notice = st.session_state.get("student_notice")
    if s_notice and s_notice.get("student_id") == student_id:
        kind = s_notice.get("kind", "info")
        msg = s_notice.get("msg", "")
        if kind == "success":
            st.success(msg)
        elif kind == "error":
            st.error(msg)
        elif kind == "warning":
            st.warning(msg)
        else:
            st.info(msg)

    show_subjects = ALL_SUBJECTS if subj_filter == "Tất cả" else [
        s for s in ALL_SUBJECTS if DISPLAY_NAME.get(s, s) == subj_filter
    ]

    shown = 0
    for subj in show_subjects:
        qs = get_quizzes(subj)
        if only_mand:
            qs = [q for q in qs if q.get("mandatory")]
        if not qs:
            continue

        qs_sorted = sorted(qs, key=lambda q: (not bool(q.get("mandatory")), safe_str(q.get("due_at") or ""), safe_str(q.get("created_at") or "")))

        st.markdown(f"### {DISPLAY_NAME.get(subj, subj)}")

        for q in qs_sorted:
            qid = q["id"]
            badge = quiz_badge(q)
            title = q.get("title", "(No title)")
            qtype = q.get("type", "mcq")
            diffq = int(q.get("difficulty", 3))
            pts = int(q.get("points", 10))
            maxa = int(q.get("max_attempts", 1))
            due_dt = parse_due(q.get("due_at"))
            due_txt = format_due(due_dt)
            maxa_txt = "∞" if maxa <= 0 else str(maxa)

            used = attempts_used(student_id, qid)
            left = attempts_left(student_id, q)
            bs = best_score(student_id, qid)
            correct_flag = any_correct(student_id, qid)
            last = last_attempt(student_id, qid)

            status = f" | attempts {used}/{maxa_txt} | best {bs}đ"
            status += " | ✅ đã đúng" if correct_flag else ""
            status += " | 🕒 overdue" if (due_dt and datetime.now() > due_dt and used == 0 and q.get("mandatory")) else ""

            with st.expander(f"{badge} — ⭐{diffq}/5 — {pts}đ — due: {due_txt} — {title}{status}"):
                st.markdown(f"**Câu hỏi:** {q.get('question','')}")
                st.caption(f"Points: {pts} | Max attempts: {maxa_txt} | Còn lượt: {left} | Deadline: {due_txt}")

                ans_key = f"ans_{student_id}_{qid}"

                if qtype == "mcq":
                    options = q.get("options", {}) or {}
                    letters = [L for L in MCQ_LETTERS if options.get(L)]
                    if letters:
                        st.radio(
                            "Chọn đáp án",
                            options=letters,
                            format_func=lambda L: f"{L}. {options.get(L,'')}",
                            key=ans_key,
                        )
                    else:
                        st.warning("Quiz này thiếu options (giáo viên cần chỉnh).")
                else:
                    st.text_area("Nhập câu trả lời", key=ans_key, height=110)

                b1, b2 = st.columns([1, 1])
                with b1:
                    st.button(
                        "📨 Nộp bài",
                        key=f"submit_{student_id}_{qid}",
                        on_click=grade_quiz,
                        args=(student_id, subj, qid),
                        use_container_width=True,
                    )
                with b2:
                    if last:
                        st.write(f"Last: {'✅ Đúng' if last.get('correct') else '❌ Sai'} | +{last.get('score',0)}đ"
                                 + (" | NỘP MUỘN" if last.get("late") else ""))
                        st.caption(f"Nộp lúc: {last.get('submitted_at')}")

                # attempt history
                atts = get_attempts(student_id, qid)
                if atts:
                    with st.expander("Lịch sử attempts"):
                        for i_att, a in enumerate(atts, start=1):
                            st.write(
                                f"{i_att}) {'✅' if a.get('correct') else '❌'} "
                                f"+{a.get('score',0)}đ"
                                + (" | NỘP MUỘN" if a.get("late") else "")
                                + f" | {a.get('submitted_at')}"
                            )

            shown += 1

    if shown == 0:
        st.info("Chưa có quiz nào (hoặc filter đang ẩn hết).")

    st.markdown("</div>", unsafe_allow_html=True)


# RIGHT: Todo + Mandatory list
with right_col:
    st.markdown('<div class="rightPanel">', unsafe_allow_html=True)
    st.markdown('<div class="todo-title">Cần làm</div>', unsafe_allow_html=True)

    def todo(icon, title, course, due, points=None):
        pts = f"<small><b>{points} điểm</b></small><br>" if points is not None else ""
        st.markdown(
            f"""
            <div class="todo-item">
              <div>{icon}</div>
              <div>
                <b>{title}</b>
                <small>{course}</small>
                {pts}
                <small>{due}</small>
              </div>
              <div class="todo-x">×</div>
            </div>
            """,
            unsafe_allow_html=True
        )

    # static demo todo
    todo("💬", "4. Hướng dẫn về nhà tuần 22", "geography", "ThO2 13 tại 11:59")
    todo("📝", "4. NỘP BÀI THI GIỮA KÌ 2 TẠI ĐÂY - DEADLINE 13/2/2026", "geography", "ThO2 13 tại 23:59", points=10)

    st.markdown("---")
    st.markdown('<div class="todo-title">Quiz bắt buộc (theo học sinh)</div>', unsafe_allow_html=True)

    student_id = st.session_state.get("current_student", {}).get("ID", None)
    if student_id:
        pending = get_pending_mandatory_quizzes(student_id)
        if not pending:
            st.success("Không còn quiz bắt buộc nào chưa làm ✅")
        else:
            now = datetime.now()
            for subj, q in pending[:8]:
                title = q.get("title", "(No title)")
                pts = int(q.get("points", 10))
                due_dt = parse_due(q.get("due_at"))
                overdue = bool(due_dt and now > due_dt)
                due_txt = format_due(due_dt)
                due_show = ("🕒 QUÁ HẠN — " if overdue else "") + due_txt
                todo("🧩", title, DISPLAY_NAME.get(subj, subj), due_show, points=pts)
            if len(pending) > 8:
                st.caption(f"Còn {len(pending)-8} quiz bắt buộc nữa...")
    else:
        st.info("Chọn học sinh để xem quiz bắt buộc.")

    st.markdown("</div>", unsafe_allow_html=True)


# =========================
# FLOATING AI CHAT
# =========================
chat_container = st.container()

with chat_container:
    st_float_container(
        key="ai-chat",
        css="""
          position: fixed !important;
          right: 20px !important;
          bottom: 20px !important;
          width: min(380px, 95vw) !important;
          max-height: 85vh !important;
        """
    )

    if st.session_state.chat_minimized:
        pill_cols = st.columns([3, 1])
        with pill_cols[0]:
            st.markdown(
                """
                <div class="chat-pill">
                  <div class="dot"></div>
                  <div>AI Assistant</div>
                </div>
                """,
                unsafe_allow_html=True,
            )
        with pill_cols[1]:
            st.markdown('<div class="miniBtn">', unsafe_allow_html=True)
            st.button("Mở", on_click=toggle_minimize, use_container_width=True)
            st.markdown("</div>", unsafe_allow_html=True)

    else:
        bubbles = []
        for msg in st.session_state.chat_history:
            role = msg.get("role", "assistant")
            content = html_safe(msg.get("content", ""))
            if role == "user":
                bubbles.append(f'<div class="bubble user"><div class="bname">Bạn</div>{content}</div>')
            else:
                bubbles.append(f'<div class="bubble ai"><div class="bname">AI</div>{content}</div>')

        st.markdown(
            f"""
            <div class="chat-shell" style="display:flex; flex-direction:column; height:min(520px, 80vh);">
              <div class="chat-head">
                <div>AI Assistant</div>
                <div class="chat-hint">Gõ: “Cho mình quiz môn Toán”</div>
              </div>

              <div class="chat-body" id="{st.session_state.chat_body_id}" style="flex:1; overflow-y:auto;">
                {''.join(bubbles)}
              </div>

              <div class="chat-foot">
            """,
            unsafe_allow_html=True
        )

        st.text_input(
            "AI",
            key="ai_question",
            placeholder="Nhập câu hỏi…",
            label_visibility="collapsed",
        )

        a, b, c = st.columns([1, 1, 1])
        with a:
            st.button("Gửi", on_click=send_ai_message, use_container_width=True)
        with b:
            st.button("Xoá", on_click=clear_ai_chat, use_container_width=True)
        with c:
            st.button("Thu nhỏ", on_click=toggle_minimize, use_container_width=True)

        st.markdown("</div></div>", unsafe_allow_html=True)

        if st.session_state.chat_autoscroll:
            components.html(
                f"""
                <script>
                  (function(){{
                    const el = window.parent.document.getElementById("{st.session_state.chat_body_id}");
                    if (el) el.scrollTop = el.scrollHeight;
                  }})();
                </script>
                """,
                height=0,
                width=0,
            )
            st.session_state.chat_autoscroll = False
